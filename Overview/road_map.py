import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

# Create a new workbook
wb = openpyxl.Workbook()

# Function to style headers: bold, centered, light blue background
def style_header(row, sheet, columns):
    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, columns + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

# Function to add borders to data rows
def add_borders(sheet, start_row, end_row, columns):
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in range(start_row, end_row + 1):
        for col in range(1, columns + 1):
            sheet.cell(row=row, column=col).border = border

# Common headers for each roadmap sheet
headers_roadmap = ["Stage/Level", "Topics/Skills", "Resources", "Projects/Practice", "Estimated Duration"]

# Helper function to create a roadmap sheet
def create_roadmap_sheet(sheet_name, roadmap_data):
    ws = wb.create_sheet(sheet_name)
    for col, header in enumerate(headers_roadmap, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header(1, ws, len(headers_roadmap))
    
    for row_idx, row_data in enumerate(roadmap_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    add_borders(ws, 2, len(roadmap_data) + 1, len(headers_roadmap))

# Roadmap 1: ERP Development => BC/NAV
bc_nav_data = [
    ["Beginner", "Learn AL language basics, Business Central fundamentals, tables, pages, codeunits", "Microsoft Learn: Dynamics 365 Business Central developer path; MB-800 study guide", "Set up a sandbox environment, create a simple table extension", "2-4 weeks"],
    ["Intermediate", "Extensions development, RDLC reports, API integration, data migration", "Microsoft Docs: 2025 release wave 1 & 2 features (AI, reporting); Certification prep for MB-800", "Build custom financial module, automate loan schedules", "4-6 weeks"],
    ["Advanced", "Power Platform integration, AI Copilot features, performance optimization, DevOps with Git", "Microsoft Learn advanced modules; 2025 wave updates on AI and compliance", "Large-scale migration project, integrate with Power Apps", "6-8 weeks"],
    ["Expert", "Custom solutions for critical sectors, troubleshooting production issues, community contributions", "GitHub repos, forums; Key features from 2025 release wave 2", "Open-source BC extension, contribute to Microsoft roadmap", "Ongoing, 3+ months"]
]

create_roadmap_sheet("BC_NAV_Roadmap", bc_nav_data)

# Roadmap 2: CRM Development => Power Platform
power_platform_data = [
    ["Beginner", "Power Apps fundamentals, canvas/model-driven apps, Dataverse basics", "Microsoft Learn: PL-900 Fundamentals; Power Apps Developer Plan (free environment)", "Build a simple canvas app for task tracking", "2-3 weeks"],
    ["Intermediate", "Power Automate flows, connectors, AI Builder, expressions", "PL-400 Developer Associate prep; 2025 release wave 1 features (Copilot Studio)", "Automate workflows, integrate with Dynamics 365", "4-6 weeks"],
    ["Advanced", "Custom components, PCF controls, security roles, advanced AI integration", "Microsoft Docs: 2025 wave 2 updates (AI-powered tools); Community forums", "Develop custom connector, AI-driven app", "6-8 weeks"],
    ["Expert", "Enterprise-scale solutions, governance, performance tuning, certification mastery", "Power Platform conferences, blogs; Updates from June 2025 feature release", "Full CRM system with Power Pages, contribute to open-source", "Ongoing, 3+ months"]
]

create_roadmap_sheet("Power_Platform_Roadmap", power_platform_data)

# Roadmap 3: C# and ASP.NET
csharp_aspnet_data = [
    ["Beginner", "C# basics (variables, loops, OOP), .NET Core essentials", "Microsoft Learn: C# beginner path; .NET Developer Roadmap on GitHub", "Simple console app, basic calculator", "2-4 weeks"],
    ["Intermediate", "ASP.NET Core MVC, Razor Pages, Entity Framework, REST APIs", "roadmap.sh/aspnet-core; Udemy/Pluralsight courses", "Build a CRUD web app with database", "4-6 weeks"],
    ["Advanced", "Authentication/Security (JWT, OAuth), Blazor, Microservices, Docker", ".NET 8/9 updates for 2025; Advanced GitHub roadmaps", "Secure API with auth, deploy to Azure", "6-8 weeks"],
    ["Expert", "Performance optimization, CI/CD, cloud integration (Azure/AWS), contributions", "YouTube: Brutally Honest Advice for .NET Devs; Community projects", "Full-stack enterprise app, open-source contribution", "Ongoing, 3+ months"]
]

create_roadmap_sheet("CSharp_ASPNET_Roadmap", csharp_aspnet_data)

# Roadmap 4: Mobile Development => Flutter
flutter_data = [
    ["Beginner", "Dart language basics, Flutter setup, widgets, layout", "Flutter.dev official docs; roadmap.sh/flutter", "Simple UI app (counter, list view)", "2-4 weeks"],
    ["Intermediate", "State management (Provider, Riverpod), Navigation, Animations, HTTP requests", "Flutter 2025 roadmap (performance, AI); GitHub flutter-development-roadmap", "To-do list app with local storage", "4-6 weeks"],
    ["Advanced", "Firebase integration, Custom widgets, Testing, Platform channels", "Medium: Flutter 2025 updates (AI, accessibility); Udemy complete guide", "E-commerce app with auth and backend", "6-8 weeks"],
    ["Expert", "Performance optimization, Web/Desktop support, AI/ML integration, contributions", "Flutter team roadmap Q1-Q4 2025; Community repos", "Cross-platform enterprise app, contribute to Flutter GitHub", "Ongoing, 3+ months"]
]

create_roadmap_sheet("Flutter_Roadmap", flutter_data)

# Roadmap 5: Master Python => DSA
python_dsa_data = [
    ["Beginner", "Python basics (syntax, functions, OOP), Arrays, Strings", "GeeksforGeeks DSA roadmap; Programiz PRO DSA with Python", "Simple scripts, array manipulations", "2-4 weeks"],
    ["Intermediate", "Linked Lists, Stacks, Queues, Recursion, Sorting (Bubble, Quick)", "roadmap.sh/datastructures-and-algorithms; LeetCode easy problems", "Implement stack/queue, sort algorithms", "4-6 weeks"],
    ["Advanced", "Trees (BST, AVL), Graphs (DFS, BFS), Dynamic Programming, Greedy", "DataCamp DSA roadmap; HackerRank practice", "LeetCode medium problems, graph traversals", "6-8 weeks"],
    ["Expert", "Advanced topics (Trie, Segment Trees), Optimization, System Design basics", "Medium: Perfect Roadmap for DSA 2025; 200+ LeetCode problems", "Full DSA projects, competitive coding contests", "Ongoing, 3+ months"]
]

create_roadmap_sheet("Python_DSA_Roadmap", python_dsa_data)

# Roadmap 6: Master MERN => Project Development
mern_data = [
    ["Beginner", "HTML/CSS/JS basics, MongoDB setup, Express/Node intro", "DEV Community MERN 2025 roadmap; roadmap.sh/full-stack", "Simple Node server, MongoDB CRUD", "2-4 weeks"],
    ["Intermediate", "React (components, hooks, state), Express APIs, Authentication", "GitHub AspNetCore-Developer-Roadmap (adapted for MERN); Medium MERN roadmap", "To-do app with React frontend", "4-6 weeks"],
    ["Advanced", "Full MERN integration, Redux, Deployment (Heroku/Vercel), Testing", "LinkedIn Ultimate MERN 2025; Projects like blog CMS", "E-commerce site with auth and payments", "6-8 weeks"],
    ["Expert", "Optimization, Scalability, TypeScript, Microservices, contributions", "Reddit MERN high-quality jobs roadmap; Open-source projects", "Enterprise full-stack app, portfolio deployment", "Ongoing, 3+ months"]
]

create_roadmap_sheet("MERN_Roadmap", mern_data)

# Adjust column widths for readability
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Save the workbook
wb.save("Roadmaps_To_Expertise.xlsx")
print("Excel file with roadmaps generated successfully!")