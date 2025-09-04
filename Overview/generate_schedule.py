import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta

# Create a new workbook
wb = openpyxl.Workbook()

# Color palette for skills (hex codes for distinct, theme-friendly colors)
skill_colors = {
    "BC/NAV": "4B8BBE",  # Blue
    "Power Platform": "2ECC71",  # Green
    "C#/ASP.NET": "9B59B6",  # Purple
    "Flutter": "3498DB",  # Light Blue
    "Python DSA": "E74C3C",  # Red
    "MERN": "F1C40F",  # Yellow
    "Exercise": "95A5A6",  # Gray
    "-": "FFFFFF"  # White for non-skill tasks
}

# Function to style headers: bold, centered, gradient blue
def style_header(row, sheet, columns, color="4FC3F7"):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, columns + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

# Function to style data rows with skill-specific colors
def style_data_row(sheet, row, skill, priority):
    fill = PatternFill(start_color=skill_colors.get(skill, "FFFFFF"), end_color=skill_colors.get(skill, "FFFFFF"), fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font = Font(bold=True if priority == "H" else False)
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill
        cell.border = border
        cell.font = font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Tab 1: Daily Schedule
ws_daily = wb.active
ws_daily.title = "Daily Schedule"
headers_daily = ["Time", "Activity", "Skill Focus", "Priority", "Details", "Motivation"]
for col, header in enumerate(headers_daily, start=1):
    ws_daily.cell(row=1, column=col, value=header)
style_header(1, ws_daily, len(headers_daily))

# Sample Monday data (Month 1, Sep 2025)
daily_data = [
    ["4:00 AM", "Wake Up", "-", "L", "Hydrate, 5-min stretch", "Your marathon starts now!"],
    ["4:15 AM", "Exercise", "Exercise", "M", "45-min run (Strava, 10k steps)", "Fitness fuels your coding grind."],
    ["5:00 AM", "Study", "BC/NAV", "H", "Microsoft Learn: MB-800 AL coding (tables, pages)", "This cert will hit 50k+ KSH."],
    ["6:00 AM", "Breakfast", "-", "L", "Oats, eggs, coffee for energy", "Fuel to outwork peers."],
    ["6:30 AM", "Study", "BC/NAV", "H", "Build BC extension (e.g., customer report in AL)", "Projects make you irreplaceable."],
    ["7:00 AM", "Study", "Power Platform", "M", "Microsoft Learn: PL-400 Power Apps canvas app", "Automation opens global roles."],
    ["7:30 AM", "Commute/Prep", "-", "L", "Podcast: ‘The Changelog’ (dev trends)", "Learn on the go."],
    ["8:00 AM–5:00 PM", "Work", "BC/NAV", "H", "Apply AL skills, log wins (e.g., automation saved 5 hrs)", "Your job is your lab."],
    ["5:00 PM", "Wind Down", "-", "L", "15-min walk (counts toward exercise)", "Breaks recharge your hustle."],
    ["6:00 PM", "Dinner", "-", "L", "Protein, veggies (e.g., chicken, greens)", "Nourish to conquer."],
    ["6:30 PM", "Study", "Python DSA", "M", "LeetCode: 2 easy problems (arrays, strings)", "DSA builds your elite edge."],
    ["7:00 PM", "Study", "C#/ASP.NET", "L", "Microsoft Learn: ASP.NET Core intro", "Expand your .NET skills."],
    ["7:30 PM", "Study", "Flutter", "L", "Flutter docs: Build basic widget (e.g., button)", "Mobile skills for startups."],
    ["8:00 PM", "Study", "MERN", "L", "freeCodeCamp: React component basics", "Web dev boosts freelance rates."],
    ["8:30 PM", "Leisure/Buffer", "-", "L", "Journal progress, read non-tech", "Balance fuels your marathon."],
    ["9:00 PM", "Sleep", "-", "M", "4-5 hrs (9:00 PM–1:30/2:00 AM)", "Rest to reset for greatness."]
]

for row_idx, row_data in enumerate(daily_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_daily.cell(row=row_idx, column=col_idx, value=value)
    style_data_row(ws_daily, row_idx, row_data[2], row_data[3])

# Notes for customization
ws_daily.cell(row=len(daily_data)+3, column=1, value="Note: Mon-Fri schedule. Adjust tasks monthly. Saturday: Full rest. Sunday: Light review (1-2 hrs).")

# Tab 2: Weekly Goals
ws_weekly = wb.create_sheet("Weekly Goals")
headers_weekly = ["Week", "Date Range", "Skill", "Priority", "Goal", "Hours Planned", "Hours Completed", "Progress", "Motivation"]
for col, header in enumerate(headers_weekly, start=1):
    ws_weekly.cell(row=1, column=col, value=header)
style_header(1, ws_weekly, len(headers_weekly))

# Weekly goals template (Month 1)
weekly_data_template = [
    ["BC/NAV", "H", "Complete 3 MB-800 modules (AL, reports)", 5, "[Fill]", "[Fill]", "Certs land 50k+ KSH jobs."],
    ["Power Platform", "M", "Complete 1 PL-400 lab (Power Apps)", 3, "[Fill]", "[Fill]", "Automation skills are your edge."],
    ["Python DSA", "M", "Solve 5 LeetCode easy problems (arrays)", 2, "[Fill]", "[Fill]", "DSA unlocks elite roles."],
    ["C#/ASP.NET", "L", "Microsoft Learn: ASP.NET Core intro (1 module)", 1, "[Fill]", "[Fill]", ".NET expands your options."],
    ["Flutter", "L", "Flutter docs: Build basic widget (e.g., button)", 1, "[Fill]", "[Fill]", "Mobile skills for startups."],
    ["MERN", "L", "freeCodeCamp: React component basics", 1, "[Fill]", "[Fill]", "Web dev boosts freelance rates."],
    ["Exercise", "M", "5 sessions (45 min each, run/yoga)", 4, "[Fill]", "[Fill]", "Fitness fuels your marathon."]
]

row_idx = 2
start_date = datetime(2025, 9, 3)
for week_num in range(1, 27):
    week_start = start_date + timedelta(weeks=week_num-1)
    week_end = week_start + timedelta(days=6)
    for skill_data in weekly_data_template:
        ws_weekly.cell(row=row_idx, column=1, value=week_num)
        ws_weekly.cell(row=row_idx, column=2, value=f"{week_start.strftime('%b %d')}–{week_end.strftime('%b %d, %Y')}")
        for col_idx, value in enumerate(skill_data, start=3):
            ws_weekly.cell(row=row_idx, column=col_idx+2, value=value)
        style_data_row(ws_weekly, row_idx, skill_data[0], skill_data[1])
        row_idx += 1

# Add table style for Weekly Goals
tab = Table(displayName="WeeklyGoals", ref=f"A1:{chr(64+len(headers_weekly))}{row_idx-1}")
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws_weekly.add_table(tab)

# Tab 3: Monthly Milestones
ws_monthly = wb.create_sheet("Monthly Milestones")
headers_monthly = ["Month", "Milestone", "Skill", "Priority", "Status", "Marketability Impact", "Motivation"]
for col, header in enumerate(headers_monthly, start=1):
    ws_monthly.cell(row=1, column=col, value=header)
style_header(1, ws_monthly, len(headers_monthly))

monthly_data = [
    ["Sep 2025", "Pass MB-800 cert", "BC/NAV", "H", "Not Started", "Qualifies for BC roles at 50k–80k KSH", "This cert puts you ahead of 90% of peers."],
    ["Oct 2025", "Pass PL-400 cert", "Power Platform", "H", "Not Started", "Adds enterprise automation skills", "Global firms need your skills."],
    ["Nov 2025", "Build ASP.NET REST API", "C#/ASP.NET", "M", "Not Started", "Expands .NET job options (40k–70k KSH)", "Versatility makes you unstoppable."],
    ["Dec 2025", "Deploy Flutter to-do app", "Flutter", "M", "Not Started", "Opens mobile dev roles (40k–70k KSH)", "Mobile skills are your startup ticket."],
    ["Jan 2026", "Solve 200+ DSA problems", "Python DSA", "H", "Not Started", "Nails tech interviews", "DSA mastery lands elite roles."],
    ["Feb 2026", "Deploy 2 MERN projects", "MERN", "H", "Not Started", "Boosts freelance rates ($20–40/hr)", "Your portfolio will scream ‘hire me’!"]
]

for row_idx, row_data in enumerate(monthly_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_monthly.cell(row=row_idx, column=col_idx, value=value)
    style_data_row(ws_monthly, row_idx, row_data[2], row_data[3])

# Tab 4: Motivational Quotes (expanded to 31 days)
ws_quotes = wb.create_sheet("Motivational Quotes")
headers_quotes = ["Day", "Quote", "Why It Matters"]
for col, header in enumerate(headers_quotes, start=1):
    ws_quotes.cell(row=1, column=col, value=header)
style_header(1, ws_quotes, len(headers_quotes))

quotes_data = [
    [1, "The only limit is the one you set. – Elon Musk", "Your 15+ hr/week study outpaces peers."],
    [2, "Discipline is freedom.", "Your daily grind builds a top-tier resume."],
    [3, "You don’t need more time, you need focus.", "Your hustle will land global gigs."],
    [4, "Grind now, shine later.", "Every LeetCode problem builds your edge."],
    [5, "Success is small efforts, repeated daily.", "Your certs will land 40k+ KSH roles."],
    [6, "Work hard in silence; let success make the noise.", "Your portfolio will speak for itself."],
    [7, "The future belongs to those who prepare today.", "Global opportunities await your skills."],
    [8, "Discipline turns dreams into reality.", "Your hustle will hit 50k KSH by Feb 2026."],
    [9, "You’re not tired, you’re uninspired—keep going!", "Your projects will wow employers."],
    [10, "The harder you work, the luckier you get.", "Your BC cert opens big doors."],
    [11, "Stay hungry, stay foolish. – Steve Jobs", "Innovate with every Power App."],
    [12, "Every expert was once a beginner.", "Your Flutter app will shine."],
    [13, "Push yourself, because no one else will.", "DSA mastery sets you apart."],
    [14, "Dream big, work hard, stay focused.", "MERN projects boost your portfolio."],
    [15, "You’re one step closer every day.", "Your skills are your ticket to 60k KSH."],
    [16, "The pain of discipline beats the pain of regret.", "Your hustle will pay off."],
    [17, "Build your future, one code at a time.", "Your portfolio will impress globally."],
    [18, "Don’t wait for opportunity, create it.", "Your certs make you a top pick."],
    [19, "Your only competition is yesterday’s you.", "Every task moves you forward."],
    [20, "Great things come from hard work.", "Your projects will land freelance gigs."],
    [21, "The best time to start was yesterday. The next best is now.", "Your skills are in demand."],
    [22, "Turn your dreams into plans.", "Your plan will hit 40k+ KSH."],
    [23, "You’re building a legacy, not just code.", "Your portfolio will stand out."],
    [24, "Keep going, you’re almost there.", "Your certs are your career rocket."],
    [25, "Success is a marathon, not a sprint.", "Your 6-month grind will pay off."],
    [26, "You’re stronger than your excuses.", "Your skills will open global doors."],
    [27, "Every hour invested builds your future.", "Your portfolio will scream value."],
    [28, "Don’t stop when it’s hard, stop when it’s done.", "Your hustle will land big roles."],
    [29, "Your work ethic is your superpower.", "Your certs will make you elite."],
    [30, "The grind is temporary, the results are permanent.", "Your skills will shine by Feb 2026."],
    [31, "You’re unstoppable when you commit.", "Your portfolio will land 50k+ KSH jobs."]
]

for row_idx, row_data in enumerate(quotes_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_quotes.cell(row=row_idx, column=col_idx, value=value)
    style_data_row(ws_quotes, row_idx, "-", "L")

# Tab 5: Progress Dashboard
ws_dashboard = wb.create_sheet("Progress Dashboard")
headers_dashboard = ["Skill", "Total Hours Planned", "Total Hours Completed", "Progress (%)", "Status"]
for col, header in enumerate(headers_dashboard, start=1):
    ws_dashboard.cell(row=1, column=col, value=header)
style_header(1, ws_dashboard, len(headers_dashboard), color="2196F3")

dashboard_data = [
    ["BC/NAV", 150, "[Fill]", "=C2/B2*100", "[Fill]"],
    ["Power Platform", 120, "[Fill]", "=C3/B3*100", "[Fill]"],
    ["C#/ASP.NET", 80, "[Fill]", "=C4/B4*100", "[Fill]"],
    ["Flutter", 80, "[Fill]", "=C5/B5*100", "[Fill]"],
    ["Python DSA", 100, "[Fill]", "=C6/B6*100", "[Fill]"],
    ["MERN", 80, "[Fill]", "=C7/B7*100", "[Fill]"]
]

for row_idx, row_data in enumerate(dashboard_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_dashboard.cell(row=row_idx, column=col_idx, value=value)
    style_data_row(ws_dashboard, row_idx, row_data[0], "M")

# Adjust column widths and row heights
for sheet in wb:
    sheet.row_dimensions[1].height = 30
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        sheet.column_dimensions[column].width = adjusted_width

# Save the workbook
wb.save("6-Month_Marathon_Upskilling_Plan_Enhanced.xlsx")
print("Enhanced Excel file generated successfully!")